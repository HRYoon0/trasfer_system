#!/usr/bin/env python3
"""통계표 헤더 전체 분석"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path, data_only=True)  # 수식 대신 값으로

ws = wb['통계표']

print("=== 통계표 헤더 (3~5행) ===\n")

# 3행 헤더
print("[3행 헤더]")
for col in range(1, 83):
    cell = ws.cell(row=3, column=col)
    if cell.value:
        from openpyxl.utils import get_column_letter
        print(f"  {get_column_letter(col)}3: {cell.value}")

# 4행 헤더
print("\n[4행 헤더]")
for col in range(1, 83):
    cell = ws.cell(row=4, column=col)
    if cell.value:
        from openpyxl.utils import get_column_letter
        print(f"  {get_column_letter(col)}4: {cell.value}")

# 5행 헤더
print("\n[5행 헤더]")
for col in range(1, 83):
    cell = ws.cell(row=5, column=col)
    if cell.value:
        from openpyxl.utils import get_column_letter
        print(f"  {get_column_letter(col)}5: {cell.value}")
