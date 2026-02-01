#!/usr/bin/env python3
"""원본 학교별현황 시트를 템플릿으로 추출"""

from openpyxl import load_workbook, Workbook
from copy import copy
import os

# 원본 파일 로드
source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb_source = load_workbook(source_path)
ws_source = wb_source['학교별현황']

# 새 워크북 생성
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '학교별현황'

# 열 너비 복사
for i, col_dim in ws_source.column_dimensions.items():
    ws_new.column_dimensions[i].width = col_dim.width

# 행 높이 복사
for i, row_dim in ws_source.row_dimensions.items():
    ws_new.row_dimensions[i].height = row_dim.height

# 플레이스홀더 매핑
placeholders = {
    'B3': '{{학교명}}',
    'B7': '{{학교명_full}}',
    'B10': '{{정원}}',
    'C10': '{{현원}}',
    'D10': '{{결원}}',
    'E10': '{{충원}}',
    'F10': '{{전출}}',
    'G10': '{{전입}}',
    'H10': '{{과부족}}',
    'D13': '{{결원명단}}',
    'H13': '{{결원계}}',
    'D14': '{{충원명단}}',
    'H14': '{{충원계}}',
    'D15': '{{관내전출명단}}',
    'H15': '{{관내전출계}}',
    'D16': '{{관외전출명단}}',
    'H16': '{{관외전출계}}',
    'D17': '{{관내전입명단}}',
    'H17': '{{관내전입계}}',
    'D18': '{{관외전입명단}}',
    'H18': '{{관외전입계}}',
}

# 셀 복사 (값, 스타일, 병합 등)
for row in ws_source.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
    for cell in row:
        new_cell = ws_new.cell(row=cell.row, column=cell.column)

        # 값 복사 (플레이스홀더 적용)
        if cell.coordinate in placeholders:
            new_cell.value = placeholders[cell.coordinate]
        else:
            new_cell.value = cell.value

        # 스타일 복사
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

# 병합 셀 복사
for merged_range in ws_source.merged_cells.ranges:
    if merged_range.min_row <= 20:
        ws_new.merge_cells(str(merged_range))

# 프론트엔드 public 폴더에 저장
output_path = '/Users/hr/Desktop/cc project/Transfer_master/frontend/public/templates/school_status_template.xlsx'
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb_new.save(output_path)
print(f"템플릿 저장 완료: {output_path}")
