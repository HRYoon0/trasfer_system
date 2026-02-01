#!/usr/bin/env python3
"""발령대장 템플릿 전체 추출 (행 제한 없이)"""

from openpyxl import load_workbook, Workbook
from copy import copy
import os

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
output_dir = '/Users/hr/Desktop/cc project/Transfer_master/frontend/public/templates'

wb_source = load_workbook(source_path)
ws_source = wb_source['발령대장']

print(f"원본 발령대장 시트: 행={ws_source.max_row}, 열={ws_source.max_column}")

# 새 워크북 생성
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '발령대장'

# 열 너비 복사
for col_letter, col_dim in ws_source.column_dimensions.items():
    ws_new.column_dimensions[col_letter].width = col_dim.width
    ws_new.column_dimensions[col_letter].hidden = col_dim.hidden

# 행 높이 복사
for row_num, row_dim in ws_source.row_dimensions.items():
    ws_new.row_dimensions[row_num].height = row_dim.height
    ws_new.row_dimensions[row_num].hidden = row_dim.hidden

# 셀 복사 (전체)
max_row = ws_source.max_row
max_col = ws_source.max_column

for row in ws_source.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        new_cell = ws_new.cell(row=cell.row, column=cell.column)
        new_cell.value = cell.value

        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

# 병합 셀 복사
for merged_range in ws_source.merged_cells.ranges:
    try:
        ws_new.merge_cells(str(merged_range))
    except:
        pass

# 인쇄 설정 복사
try:
    ws_new.print_title_rows = ws_source.print_title_rows
    ws_new.page_setup.orientation = ws_source.page_setup.orientation
    ws_new.page_setup.paperSize = ws_source.page_setup.paperSize
    ws_new.page_margins = copy(ws_source.page_margins)
except:
    pass

# 저장
output_path = os.path.join(output_dir, 'assignment_list_template.xlsx')
wb_new.save(output_path)
print(f"저장 완료: {output_path}")
print(f"추출된 행: {max_row}, 열: {max_col}")
