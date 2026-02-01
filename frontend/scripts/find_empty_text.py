#!/usr/bin/env python3
"""발령대장에서 '이하 여백' 관련 텍스트 찾기"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path)
ws = wb['발령대장']

print("=== 발령대장에서 '여백' 또는 '이하' 텍스트 찾기 ===\n")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if '여백' in cell.value or '이하' in cell.value:
                print(f"{cell.coordinate}: '{cell.value}'")
                # 주변 셀도 확인
                print(f"  - 스타일: font={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}")
                print(f"  - 정렬: horizontal={cell.alignment.horizontal}")

print("\n\n=== 발령대장 4~10행 D열 내용 확인 ===")
for row_num in range(4, 15):
    cell = ws.cell(row=row_num, column=4)  # D열
    print(f"D{row_num}: '{cell.value}' (type: {type(cell.value).__name__})")
