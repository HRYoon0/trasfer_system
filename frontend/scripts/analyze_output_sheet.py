#!/usr/bin/env python3
"""출력물관리 시트 분석 - 버튼 및 구조 확인"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path)

# 출력물관리 시트 확인
ws = wb['출력물관리']

print("=== 출력물관리 시트 내용 ===")
print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}\n")

# 값이 있는 셀 출력
for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=15):
    for cell in row:
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value}")

# 발령대장 시트도 확인
print("\n\n=== 발령대장 시트 헤더 확인 ===")
ws2 = wb['발령대장']
for row in ws2.iter_rows(min_row=1, max_row=5, min_col=1, max_col=15):
    for cell in row:
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value}")
