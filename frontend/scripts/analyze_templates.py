#!/usr/bin/env python3
"""각 템플릿의 구조 분석"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path)

def analyze_sheet(sheet_name, max_row=30, max_col=15):
    print(f"\n{'='*60}")
    print(f"=== {sheet_name} 구조 분석 ===")
    print(f"{'='*60}")

    ws = wb[sheet_name]

    # 열 너비
    print("\n[열 너비]")
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width:
                print(f"  {col_letter}: {width:.2f}")

    # 병합 셀
    print("\n[병합 셀]")
    for merged_range in list(ws.merged_cells.ranges)[:20]:
        print(f"  {merged_range}")

    # 셀 내용 (값이 있는 셀만)
    print("\n[셀 내용]")
    for row in range(1, min(max_row + 1, ws.max_row + 1)):
        for col in range(1, min(max_col + 1, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                addr = cell.coordinate
                val = str(cell.value)[:50]
                print(f"  {addr}: {val}")

# 각 시트 분석
analyze_sheet('발령대장', 10, 15)
analyze_sheet('통지서', 20, 10)
analyze_sheet('기관통보', 10, 12)
analyze_sheet('임용서', 12, 12)
