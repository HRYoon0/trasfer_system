#!/usr/bin/env python3
"""원본 파일의 시트 목록 확인"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path, read_only=True)

print("=== 원본 파일 시트 목록 ===")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"{i}. {name}")
