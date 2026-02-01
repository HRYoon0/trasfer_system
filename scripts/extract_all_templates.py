#!/usr/bin/env python3
"""원본 파일에서 모든 문서 시트를 템플릿으로 추출"""

from openpyxl import load_workbook, Workbook
from copy import copy
import os

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
output_dir = '/Users/hr/Desktop/cc project/Transfer_master/frontend/public/templates'

os.makedirs(output_dir, exist_ok=True)

# 추출할 시트 목록
sheets_to_extract = [
    ('통지서', 'notice_template.xlsx'),
    ('통지서_타시군전출', 'notice_external_template.xlsx'),
    ('임명장', 'appointment_template.xlsx'),
    ('발령대장', 'assignment_list_template.xlsx'),
    ('기관통보', 'institution_notification_template.xlsx'),
    ('임용서', 'employment_letter_template.xlsx'),
]

wb_source = load_workbook(source_path)

for sheet_name, output_filename in sheets_to_extract:
    print(f"\n=== {sheet_name} 추출 중 ===")

    if sheet_name not in wb_source.sheetnames:
        print(f"  시트 '{sheet_name}'을 찾을 수 없습니다.")
        continue

    ws_source = wb_source[sheet_name]

    # 새 워크북 생성
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = sheet_name

    # 열 너비 복사
    for col_letter, col_dim in ws_source.column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width
        ws_new.column_dimensions[col_letter].hidden = col_dim.hidden

    # 행 높이 복사
    for row_num, row_dim in ws_source.row_dimensions.items():
        ws_new.row_dimensions[row_num].height = row_dim.height
        ws_new.row_dimensions[row_num].hidden = row_dim.hidden

    # 최대 범위 확인
    max_row = min(ws_source.max_row, 100)  # 최대 100행
    max_col = min(ws_source.max_column, 20)  # 최대 20열

    # 셀 복사
    for row in ws_source.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column)
            new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 병합 셀 복사
    for merged_range in ws_source.merged_cells.ranges:
        if merged_range.min_row <= max_row and merged_range.min_col <= max_col:
            try:
                ws_new.merge_cells(str(merged_range))
            except:
                pass

    # 인쇄 설정 복사 (가능한 경우)
    try:
        ws_new.print_title_rows = ws_source.print_title_rows
        ws_new.print_title_cols = ws_source.print_title_cols
        ws_new.page_setup.orientation = ws_source.page_setup.orientation
        ws_new.page_setup.paperSize = ws_source.page_setup.paperSize
        ws_new.page_margins = copy(ws_source.page_margins)
    except:
        pass

    # 저장
    output_path = os.path.join(output_dir, output_filename)
    wb_new.save(output_path)
    print(f"  저장 완료: {output_filename}")
    print(f"  행: {max_row}, 열: {max_col}")

print("\n=== 모든 템플릿 추출 완료 ===")
