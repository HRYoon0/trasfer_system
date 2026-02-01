#!/usr/bin/env python3
"""통계표 시트 구조 분석"""

from openpyxl import load_workbook

source_path = '/Users/hr/Desktop/cc project/Transfer_master/초등경남관내전보_작업상황부ver3.2(양산).xlsm'
wb = load_workbook(source_path)

# 통계표 시트 분석
ws = wb['통계표']

print("=== 통계표 시트 구조 ===")
print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}\n")

print("[열 너비]")
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
    if col in ws.column_dimensions and ws.column_dimensions[col].width:
        print(f"  {col}: {ws.column_dimensions[col].width:.2f}")

print("\n[병합 셀]")
for merged in list(ws.merged_cells.ranges)[:30]:
    print(f"  {merged}")

print("\n[셀 내용 (1~50행)]")
for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=15):
    for cell in row:
        if cell.value is not None:
            val = str(cell.value)[:40]
            print(f"  {cell.coordinate}: {val}")
